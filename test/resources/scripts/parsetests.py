from openpyxl import load_workbook


def open_workbook(excel_file):
    work_book  = load_workbook(filename = excel_file)
    return work_book

def index_tests(work_book):
    test_index = []
    first_sheet = work_book.worksheets[0] #Assume index is always first sheet (in excel and in openpyxl)
    first_row = 5 # Miss the header etc
    #Also assume first column is test reference
    test_ref_col = 0
    title_col = -1
    for cell in first_sheet[4]:
        if(cell.value and cell.value.strip().lower()=="test case title"):
            title_col = cell.column
    if(title_col==-1):
        raise AttributeError("Failed to find title column")

    for row in first_sheet.iter_rows(min_row = first_row):
        test_ref = row[test_ref_col].value
        test_title = row[title_col-1].value
        if test_ref==None:
            break
        test_index.append((test_ref,test_title))
    return test_index

def get_scenario(work_sheet):
    part_names = work_sheet["A"]
    for part in part_names:
        if part.value=="Scenario Outline":
            scenario_row =  part.row
            return work_sheet["B"][scenario_row-1].value

def get_givens(work_sheet):
    givens = []
    part_names = work_sheet["A"]
    in_given = False
    for part in part_names:
        if part.value.strip().lower()=="given":
            in_given = True
            given_row = part.row
            givens.append(work_sheet["B"][given_row-1].value)
        elif in_given and part.value.strip().lower()=="and":
            given_row = part.row
            givens.append(work_sheet["B"][given_row-1].value)
        elif in_given:
            break
    return givens

def get_when(work_sheet):
    part_names = work_sheet["A"]
    for part in part_names:
        if part.value=="When":
            when_row =  part.row
            return work_sheet["B"][when_row-1].value

def get_thens(work_sheet):
    thens = []
    part_names = work_sheet["A"]
    in_then = False
    for part in part_names:
        if part.value==None and not in_then:
            raise AttributeError("Failed to find a then?")
        elif part.value==None:
            break
        elif part.value.strip().lower()=="then":
            in_then = True
            then_row = part.row
            thens.append(work_sheet["B"][then_row-1].value)
        elif in_then and part.value.strip().lower()=="and":
            then_row = part.row
            thens.append(work_sheet["B"][then_row-1].value)
        elif in_then:
            break
    return thens

def make_gherkin_feature(test_ref,feature,scenario,given,when,then):
    f = open("%s.feature" % test_ref,"w")
    print("Feature: %s" % feature, file = f)
    print("", file = f)
    print("  Scenario: %s" % scenario, file = f)
    print("    Given %s" % given[0], file = f)
    for g in given[1:]:
        print("    And %s" % g, file = f)
    print("    When %s" % when, file = f) #Assuming single WHEN
    print("    Then %s" % then[0], file = f)
    for t in then[1:]:
        print("    And %s" % t, file = f)
    f.close()

if __name__=="__main__":
    #work_book = open_workbook('../Supplier Test Cases/LIMS system test cases  - consuming requests & sending reports.xlsx')
    work_book = open_workbook('../Supplier Test Cases/GP system test cases - REQUESTS.xlsx')
    test_index = index_tests(work_book)
    for test_ref,test_title in test_index:
        try:
            ws = work_book[test_ref]
        except KeyError:
            ws = work_book['%s - Optional Test' % test_ref] #Quick workaround for test in LIMS set
        scenario = get_scenario(ws)
        given = get_givens(ws)
        when = get_when(ws)
        then = get_thens(ws)
        make_gherkin_feature(test_ref,test_title,scenario,given,when,then)
        break
